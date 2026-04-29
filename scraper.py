#!/usr/bin/env python3
import os
import smtplib
from datetime import datetime, timedelta
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email import encoders
from typing import List, Dict

from jobspy import scrape_jobs
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import (
    JOB_ROLES, LOCATIONS, DAYS_BACK,
    EMAIL_SENDER, EMAIL_PASSWORD, EMAIL_RECIPIENT,
    OUTPUT_DIR, OUTPUT_FILENAME_PREFIX, RESULTS_PER_ROLE,
    EMAIL_SUBJECT, SCRAPE_LINKEDIN, SCRAPE_INDEED,
    SCRAPE_GLASSDOOR, SCRAPE_ZIPRECRUITER, SCRAPE_GOOGLE,
    TITLE_KEYWORDS,
)

# Columns to include in the Excel sheet (in order)
EXCEL_COLUMNS = [
    ("title",           "Job Title"),
    ("company",         "Company"),
    ("location",        "Location"),
    ("job_type",        "Job Type"),
    ("min_amount",      "Salary Min"),
    ("max_amount",      "Salary Max"),
    ("currency",        "Currency"),
    ("date_posted",     "Date Posted"),
    ("site",            "Source"),
    ("job_url",         "Apply Link"),
]


def get_job_boards():
    boards = []
    if SCRAPE_LINKEDIN:    boards.append("linkedin")
    if SCRAPE_INDEED:      boards.append("indeed")
    if SCRAPE_GLASSDOOR:   boards.append("glassdoor")
    if SCRAPE_ZIPRECRUITER: boards.append("zip_recruiter")
    if SCRAPE_GOOGLE:      boards.append("google")
    return boards or ["indeed", "linkedin", "glassdoor"]


def is_relevant_job(job: Dict) -> bool:
    title = str(job.get('title', '')).lower()
    return any(kw.lower() in title for kw in TITLE_KEYWORDS)


def scrape_all_jobs() -> List[Dict]:
    all_jobs = []
    print(f"🔍 Scraping {len(JOB_ROLES)} roles across {', '.join(get_job_boards())}\n")

    for role in JOB_ROLES:
        for location in LOCATIONS:
            try:
                print(f"   → {role} in {location}...")
                jobs = scrape_jobs(
                    site_name=get_job_boards(),
                    search_term=role,
                    location=location,
                    results_wanted=RESULTS_PER_ROLE,
                    hours_old=DAYS_BACK * 24,
                    country_indeed="USA",
                )
                if hasattr(jobs, 'to_dict'):
                    jobs = jobs.to_dict('records')

                before = len(jobs)
                jobs = [j for j in jobs if is_relevant_job(j)]
                all_jobs.extend(jobs)
                print(f"      ✓ {len(jobs)} relevant ({before - len(jobs)} filtered out)")

            except Exception as e:
                print(f"      ✗ Error: {str(e)}")
                continue

    # Deduplicate by title + company
    seen = set()
    unique = []
    for job in all_jobs:
        key = (str(job.get('title', '')).lower(), str(job.get('company', '')).lower())
        if key not in seen:
            seen.add(key)
            unique.append(job)

    print(f"\n✅ Total unique relevant jobs: {len(unique)}\n")
    return unique


def save_to_excel(jobs: List[Dict]) -> str:
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    today = datetime.now().strftime("%Y-%m-%d")
    filename = f"{OUTPUT_DIR}/{OUTPUT_FILENAME_PREFIX}-{today}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cybersecurity Jobs"

    # --- Styles ---
    header_fill   = PatternFill("solid", fgColor="1F4E79")
    header_font   = Font(bold=True, color="FFFFFF", size=11)
    link_font     = Font(color="0563C1", underline="single")
    alt_fill      = PatternFill("solid", fgColor="EBF3FB")
    border_side   = Side(style="thin", color="CCCCCC")
    cell_border   = Border(
        left=border_side, right=border_side,
        top=border_side,  bottom=border_side
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # --- Header row ---
    headers = [col[1] for col in EXCEL_COLUMNS]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = cell_border

    # --- Data rows ---
    for row_idx, job in enumerate(jobs, 2):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill()
        for col_idx, (key, _) in enumerate(EXCEL_COLUMNS, 1):
            value = job.get(key, "")
            if value is None or str(value) in ("nan", "NaT", "None"):
                value = ""

            cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value != "" else "")
            cell.border    = cell_border
            cell.fill      = fill

            # Make Apply Link clickable
            if key == "job_url" and value:
                cell.value     = "Apply Now"
                cell.hyperlink = str(value)
                cell.font      = link_font
                cell.alignment = center
            else:
                cell.alignment = left

    # --- Column widths ---
    col_widths = {
        "Job Title":    35,
        "Company":      25,
        "Location":     22,
        "Job Type":     12,
        "Salary Min":   12,
        "Salary Max":   12,
        "Currency":     10,
        "Date Posted":  14,
        "Source":       12,
        "Apply Link":   14,
    }
    for col_idx, (_, label) in enumerate(EXCEL_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(label, 15)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter on header row
    ws.auto_filter.ref = f"A1:{get_column_letter(len(EXCEL_COLUMNS))}1"

    wb.save(filename)
    print(f"💾 Saved {len(jobs)} jobs to Excel: {filename}")
    return filename


def send_email(jobs: List[Dict], excel_filename: str):
    sender    = os.environ.get('EMAIL_SENDER',    EMAIL_SENDER)
    password  = os.environ.get('EMAIL_PASSWORD',  EMAIL_PASSWORD)
    recipient = os.environ.get('EMAIL_RECIPIENT', EMAIL_RECIPIENT)

    print(f"📧 Sending from: {sender} → {recipient}")

    # Count by source for summary
    sources = {}
    for job in jobs:
        src = str(job.get('site', 'Unknown'))
        sources[src] = sources.get(src, 0) + 1

    source_rows = ''.join([
        f"<tr><td style='padding:8px 12px;border-bottom:1px solid #ddd;'>{s}</td>"
        f"<td style='padding:8px 12px;border-bottom:1px solid #ddd;text-align:center;'>{c}</td></tr>"
        for s, c in sorted(sources.items(), key=lambda x: x[1], reverse=True)
    ])

    # Count by role
    role_counts = {}
    for job in jobs:
        t = str(job.get('title', 'Unknown'))
        role_counts[t] = role_counts.get(t, 0) + 1
    top_roles = sorted(role_counts.items(), key=lambda x: x[1], reverse=True)[:10]
    role_rows = ''.join([
        f"<tr><td style='padding:8px 12px;border-bottom:1px solid #ddd;'>{t}</td>"
        f"<td style='padding:8px 12px;border-bottom:1px solid #ddd;text-align:center;'>{c}</td></tr>"
        for t, c in top_roles
    ])

    html = f"""
    <html><body style="font-family:Arial,sans-serif;color:#333;max-width:640px;margin:0 auto;padding:20px;">
      <h1 style="color:#1F4E79;border-bottom:3px solid #2E75B6;padding-bottom:10px;">
        Weekly Cybersecurity Jobs Report
      </h1>
      <div style="background:#EBF3FB;padding:15px;border-radius:6px;margin:16px 0;">
        <p><strong>Total Jobs Found:</strong> {len(jobs)}</p>
        <p><strong>Period:</strong> Last {DAYS_BACK} day(s)</p>
        <p><strong>Generated:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M')} UTC</p>
        <p><strong>Next Run:</strong> {(datetime.now() + timedelta(days=7)).strftime('%Y-%m-%d')}</p>
      </div>

      <h2 style="color:#1F4E79;">Top Job Titles Found</h2>
      <table style="width:100%;border-collapse:collapse;">
        <tr style="background:#1F4E79;color:#fff;">
          <th style="padding:10px 12px;text-align:left;">Job Title</th>
          <th style="padding:10px 12px;text-align:center;">Count</th>
        </tr>
        {role_rows}
      </table>

      <h2 style="color:#1F4E79;">Jobs by Source</h2>
      <table style="width:100%;border-collapse:collapse;">
        <tr style="background:#1F4E79;color:#fff;">
          <th style="padding:10px 12px;text-align:left;">Job Board</th>
          <th style="padding:10px 12px;text-align:center;">Count</th>
        </tr>
        {source_rows}
      </table>

      <div style="background:#FFF3CD;border-left:4px solid #FFC107;padding:12px;margin:20px 0;">
        <strong>Excel file attached</strong> — open it for all {len(jobs)} listings
        with clickable <em>Apply Now</em> links, salaries, locations, and dates.
      </div>

      <h2 style="color:#1F4E79;">Roles Searched</h2>
      <ul>{''.join(f'<li>{r}</li>' for r in JOB_ROLES)}</ul>

      <p style="color:#999;font-size:12px;margin-top:30px;">
        Automated weekly job scraper · Next run {(datetime.now() + timedelta(days=7)).strftime('%Y-%m-%d')}
      </p>
    </body></html>
    """

    try:
        msg = MIMEMultipart('mixed')
        msg['Subject'] = EMAIL_SUBJECT
        msg['From']    = sender
        msg['To']      = recipient
        msg.attach(MIMEText(html, 'html'))

        # Attach Excel file
        with open(excel_filename, 'rb') as f:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header(
                'Content-Disposition',
                f'attachment; filename="{os.path.basename(excel_filename)}"'
            )
            msg.attach(part)

        with smtplib.SMTP('smtp.gmail.com', 587) as server:
            server.ehlo()
            server.starttls()
            server.login(sender, password)
            server.send_message(msg)

        print(f"📧 Email + Excel attachment sent to {recipient}")

    except Exception as e:
        print(f"❌ Email error: {str(e)}")


def main():
    print("=" * 60)
    print("🚀 CYBERSECURITY JOB SCRAPER")
    print("=" * 60 + "\n")

    jobs = scrape_all_jobs()

    if not jobs:
        print("❌ No jobs found.")
        return

    excel_file = save_to_excel(jobs)
    send_email(jobs, excel_file)

    print("\n" + "=" * 60)
    print("✅ DONE!")
    print("=" * 60)


if __name__ == "__main__":
    main()
